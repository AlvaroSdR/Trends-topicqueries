import requests
import openpyxl
import datetime
import time

# Ask the user to enter a topic and the number of results they want
topic = input("Enter a topic to search for: ")
num_results = int(input("Enter the number of results you want (recommended not to exceed 10): "))

# Ask the user for the number of days to include in the search
num_days = int(input("Enter the number of days to include in the search: "))

# Calculate the dates for the specified time range
end_date = datetime.datetime.today().strftime('%Y-%m-%d')
start_date = (datetime.datetime.today() - datetime.timedelta(days=num_days)).strftime('%Y-%m-%d')

# Set the countries to search in
countries = ['US', 'CN', 'JP', 'DE', 'GB', 'IN', 'FR', 'IT', 'BR', 'CA', 'KR', 'RU', 'AU', 'ES', 'MX', 'ID', 'NL', 'SA', 'TR', 'CH', 'TW', 'PL', 'SE', 'BE', 'AR', 'TH', 'IR', 'AT', 'NO', 'AE']
countries2 = ['United States', 'China', 'Japan', 'Germany', 'United Kingdom', 'India', 'France', 'Italy', 'Brazil', 'Canada', 'South Korea', 'Russia', 'Australia', 'Spain', 'Mexico', 'Indonesia', 'Netherlands', 'Saudi Arabia', 'Turkey', 'Switzerland', 'Taiwan', 'Poland', 'Sweden', 'Belgium', 'Argentina', 'Thailand', 'Iran', 'Austria', 'Norway', 'United Arab Emirates']

# Create an Excel workbook and add a sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Top Searches'

# Write the header row
sheet.cell(row=1, column=1, value='Country')
for i in range(2, num_results + 2):
    sheet.cell(row=1, column=i, value=f'Top {i-1} Searches')

# Loop through each country
for i, country in enumerate(countries):
    # Create a payload to send to the Google Trends API
    url = 'https://api.trends.google.com/v1alpha/queries:batchGet?hl=en-US&tz=-120&'
    keyword_payload = f'q={topic}&startDate={start_date}&endDate={end_date}&geo={country}'
    keywords = keyword_payload.format(topic, start_date, end_date, country)
    payload = f'{{"time":"{start_date}T00:00:00Z/{end_date}T23:59:59Z","resolution":"COUNTRY","locale":"en-US","comparisonItem":[{{{keywords}}}]}}'

    # Send a POST request to the Google Trends API and extract the JSON data
    response = requests.post(url, data=payload)
    json_data = response.json()

    # Find the top searched queries by parsing the JSON data and storing them in a list
    searches = []
    for query in json_data['default']['timelineData'][0]['values']:
        searches.append(query['label'])

    # If the list of searches is empty, print an error message and skip to the next country
    if not searches:
        print(f"No trending searches found for {topic} in {countries2[i]} for the last {num_days} days")
        continue

    top_searches = searches[:num_results]

    # Write the country and top searches to the Excel sheet in separate columns
    sheet.cell(row=i+2, column=1, value=country)
    for j, search in enumerate(top_searches):
        sheet.cell(row=i+2, column=j+2, value=search)

# Save the Excel workbook
wb.save('top_searches.xlsx')

